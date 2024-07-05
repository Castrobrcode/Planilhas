from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from time import sleep



p = openpyxl.load_workbook('planilhas/dados_clientes.xlsx')
pagina = p['Sheet1']
driver = webdriver.Firefox()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina.iter_rows(min_row=2, values_only=True):
     nome, valor, cpf, vencimento = linha
     
     sleep(5)
     campo = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
     sleep(1)
     campo.clear()
     campo.send_keys(cpf)
     sleep(1)
     button = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
     sleep(1)
     button.click()
     sleep(4)
     status = driver.find_element(By.XPATH, '//span[@id="statusLabel"]')
     sleep(1)
     S = status.text
     print(S)
     if S == 'em dia':
        Data = driver.find_element(By.XPATH,'//p[@id="paymentDate"]')
        Data_limpo = Data.text.split()[3]
        Metodo = driver.find_element(By.XPATH,'//p[@id="paymentMethod"]')   
        Metodo_limpo = Metodo.text.split()[3]
        pf = openpyxl.load_workbook('planilhas/p.xlsx')
        pf1 = pf['Sheet1']
        pf1.append([nome, valor, cpf, vencimento, 'em dia', Data_limpo, Metodo_limpo])
        pf.save('planilhas/p.xlsx') 
     else:
         pf = openpyxl.load_workbook('planilhas/p.xlsx')
         pf1 = pf['Sheet1']         
         pf1.append([nome, valor, cpf, vencimento,'pendente'])
         pf.save('planilhas/p.xlsx')
